#!/usr/bin/env python3
"""Build a plain-language Excel report from the technical workbook.

Output/matches_report.xlsx is written for a business reader: three sheets,
plain headers, no internal scores or jargon. The technical detail stays in
Output/matches_v2.xlsx.

  1. Matches            one row per validated match, with the reason in plain
                        English and a simple confidence.
  2. No match found     opportunities where nothing in the company list fits.
  3. Needs and gaps     what each opportunity requires and who covers it.

Usage:
  python3 build_report_xlsx.py     # reads Output/matches_v2.xlsx
"""
from __future__ import annotations

import os

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SOURCE = "Output/matches_v2.xlsx"
OUT = "Output/matches_report.xlsx"
BRAND = "02714E"  # MISA green


def clean(s) -> str:
    """Plain text: no em or en dashes, no stray whitespace."""
    s = str(s if s is not None else "")
    if s == "nan":
        return ""
    return " ".join(s.replace("—", "-").replace("–", "-").split())


def confidence(row) -> str:
    if clean(row.get("human_verdict")) == "Agree":
        return "Confirmed by analyst"
    agree = clean(row.get("gpt_agreement"))
    if "/" in agree:
        k, n = agree.split("/")
        if k == n:
            return "High (unanimous)"
        return f"Medium ({agree} votes)"
    return "Medium"


def role(row) -> str:
    if row.get("gpt_decision") in ("Direct", "Yes"):
        return "Can build it"
    return "Supplier / partner"


def build_frames():
    allp = pd.read_excel(SOURCE, sheet_name="All_Pairs")
    matches = allp[allp["validated_fit"] == True].copy()  # noqa: E712
    matches = matches.sort_values(["opportunity", "final_score"],
                                  ascending=[True, False])
    m = pd.DataFrame({
        "Opportunity": matches["opportunity"].map(clean),
        "Opportunity sector": matches["opportunity_sector"].map(clean),
        "Matched company": matches["company"].map(clean),
        "Company sector": matches["company_sector"].map(clean),
        "Role": [role(r) for _, r in matches.iterrows()],
        "Why it matches": matches["gpt_explanation"].map(clean),
        "Confidence": [confidence(r) for _, r in matches.iterrows()],
        "Your verdict": matches["human_verdict"].map(clean),
    })

    ab = pd.read_excel(SOURCE, sheet_name="Abstentions")
    ab = ab[ab["opportunity"].astype(str) != "-"]
    n = pd.DataFrame({
        "Opportunity": ab["opportunity"].map(clean),
        "Finding": ["No suitable company in the current list"] * len(ab),
        "Closest candidate (not a fit)": ab.get("best_candidate", pd.Series([""] * len(ab))).map(clean),
        "Detail": ab.get("detail", pd.Series([""] * len(ab))).map(clean),
    })

    try:
        con = pd.read_excel(SOURCE, sheet_name="Consortium_View")
        con = con[con.get("need", pd.Series(dtype=str)).astype(str).str.len() > 0]
        g = pd.DataFrame({
            "Opportunity": con["opportunity"].map(clean),
            "What it needs": con["need"].map(clean),
            "Covered by": con["covered_by"].map(clean).replace("", "NOBODY YET (gap)"),
            "Notes": con.get("why", pd.Series([""] * len(con))).map(clean),
        })
    except Exception:
        g = pd.DataFrame(columns=["Opportunity", "What it needs", "Covered by", "Notes"])
    return m, n, g


WIDTHS = {
    "Opportunity": 38, "Opportunity sector": 16, "Matched company": 26,
    "Company sector": 20, "Role": 16, "Why it matches": 90,
    "Confidence": 20, "Your verdict": 12, "Finding": 34,
    "Closest candidate (not a fit)": 26, "Detail": 70,
    "What it needs": 30, "Covered by": 28, "Notes": 70,
}


def style_sheet(ws, n_rows, headers):
    head_fill = PatternFill("solid", fgColor=BRAND)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = WIDTHS.get(h, 18)
    for r in range(2, n_rows + 2):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(
                wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"


def main():
    if not os.path.exists(SOURCE):
        raise SystemExit(f"Not found: {SOURCE} (run matching_v2.py first).")
    m, n, g = build_frames()
    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        m.to_excel(writer, sheet_name="Matches", index=False)
        n.to_excel(writer, sheet_name="No match found", index=False)
        g.to_excel(writer, sheet_name="Needs and gaps", index=False)
        style_sheet(writer.sheets["Matches"], len(m), list(m.columns))
        style_sheet(writer.sheets["No match found"], len(n), list(n.columns))
        style_sheet(writer.sheets["Needs and gaps"], len(g), list(g.columns))
    print(f"Wrote {OUT}: {len(m)} matches, {len(n)} without a match, "
          f"{len(g)} need rows.")


if __name__ == "__main__":
    main()
