"""
Company–Opportunity Matching Pipeline
======================================

Implements the full 8-step architecture:

  Step 1  Preprocessing                  — clean and normalize text
  Step 2  Sector Ontology Expansion      — expand sector tags via GPT-built tree
  Step 3  Sector Filtering               — drop pairs with no sector overlap
  Step 4  Semantic Embedding & Similarity — text-embedding-ada-002 (1536-dim)
  Step 5  Product/Service Matching       — combined_score = max(profile, product)
  Step 6  GPT-Based Validation           — analyst-grade outcome for admitted pairs only
  Step 7  Soft Match Mode (Sector Relaxed)— if sectors disjoint but combined ≥ 0.85,
                                           GPT still decides → cross-sector synergies
  Step 8  Ranking and Scoring             — rank within opportunityId by final_score
                                           desc; tie-break product_similarity desc

Output: Output/business_grade_company_opportunity_matches.xlsx

Before writing the workbook, the ranked output table is saved to
Output/business_grade_pre_export.pkl — recover a failed Excel write with:

  python3 business_grade_matching.py --resume-export

After Step 6 (all GPT decisions + filtered rows filled in), the full pair
DataFrame is saved to Output/business_grade_post_step6.pkl — rerun only
ranking + export (no API) with:

  python3 business_grade_matching.py --resume-from-step8

Company source for this deployment: **kpmgfile.xlsx** (required to resolve to
that filename).

Optional **.env** in this directory: set ``OPENAI_API_KEY=...`` so runs do not
depend on a broken global shell variable (values in ``.env`` override the
environment for keys present in the file).

Export columns include human-readable source text (KPMG / opportunities file):
  company_name, opportunity_name, company_profile, opportunity_description
plus sector_similarity (semantic gate score in [0,1]), ontology_sector_overlap (0/1),
other scored fields, AI columns, and rank.

Export uses tri-state ai_decision: "Yes" (LLM accept), "No" (LLM reject),
"Filtered" (sector gate; LLM never run).

final_score = 0.30 * ai_score + 0.30 * profile_similarity + 0.40 * product_similarity

sector_similarity (output) =
  max(profile_similarity, product_similarity) — the same [0,1] gate score compared
  to SOFT_MATCH_THRESHOLD in Step 7. ontology_sector_overlap is 1 when the pair
  was admitted primarily via ontology sector-token overlap rather than semantic soft-match.
"""

import json
import math
import os
import re
import sys
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, List, Optional, Set, Tuple

import numpy as np
import pandas as pd
from openai import OpenAI
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from tqdm import tqdm

from sector_inference import closed_sector_vocabulary, enrich_companies_missing_sectors

from matcher_signals import (
    format_fail_reason,
    format_pass_reason,
    keyword_match_count,
    parse_json_list,
    resolve_adjacent_label,
)


# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
COMPANY_FILE_CANDIDATES = [
    "kpmgfile.xlsx",
    "Companies.xlsx",
    "companies.xlsx",
    "Data/companies.xlsx",
]
OPPORTUNITY_FILE_CANDIDATES = [
    "opportunities_structured.xlsx",
    "Output/opportunities_structured.xlsx",
    "new_opportunities.xlsx",
    "Data/new_opportunities.xlsx",
    "Data/Opportunitiesfinalfile.xlsx",
]
OUTPUT_FILE = "Output/business_grade_company_opportunity_matches.xlsx"
# Saved after Step 6 finishes (GPT + filtered stub rows); resume Step 8+export only.
CHECKPOINT_POST_STEP6 = "Output/business_grade_post_step6.pkl"
# Saved after Step 8 (ranked 16-col table); recover failed .xlsx with --resume-export
CHECKPOINT_PRE_EXPORT = "Output/business_grade_pre_export.pkl"
ONTOLOGY_CACHE_FILE = "Output/sector_ontology_cache.json"
SECTOR_INFERENCE_CACHE_FILE = "Output/sector_inference_cache.json"

ONTOLOGY_MODEL   = "gpt-4o-mini"           # cheap; cached
EXTRACTION_MODEL = "gpt-4o-mini"
DECISION_MODEL   = "gpt-4.1"
EMBEDDING_MODEL  = "text-embedding-ada-002"  # per Step 4 spec (1536-dim)

EXTRACT_BATCH  = 48
DECISION_BATCH = 14

# Concurrent OpenAI calls (lower MATCH_PARALLEL_* if you see HTTP 429).
PARALLEL_EMBED_CHUNKS  = max(2, min(24, int(os.getenv("MATCH_PARALLEL_EMBED", "14"))))
PARALLEL_EXTRACT_BATCH = max(2, min(32, int(os.getenv("MATCH_PARALLEL_EXTRACT", "22"))))
PARALLEL_OPP_GPT       = max(1, min(12, int(os.getenv("MATCH_PARALLEL_GPTOPP", "10"))))

ENTITY_EXTRACTION_CACHE = "Output/entity_extraction_cache.pkl"

# Legacy soft-match ceiling (shown in summaries; routing uses TRIAGE_* below.)
SOFT_MATCH_THRESHOLD = 0.85

# Multi-signal triage (Phase 2) — routed to GPT if ANY fires.
TRIAGE_MIN_CAPABILITY_KEYWORD_HITS = 2
TRIAGE_MIN_INPUT_OVERLAP_HITS = 1
TRIAGE_MIN_SIGNAL_PRODUCT_SIM = 0.80
TRIAGE_MIN_SIGNAL_PROFILE_SIM = 0.82

# final_score weights — must reproduce the reference MatchingOutput.csv:
W_AI      = 0.30
W_PROFILE = 0.30
W_PRODUCT = 0.40


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def load_project_dotenv(filename: str = ".env") -> None:
    """Load KEY=VALUE lines from script-dir .env into os.environ (file overrides env)."""
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
    if not os.path.isfile(path):
        return
    with open(path, encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            if "=" not in line:
                continue
            key, _, val = line.partition("=")
            key, val = key.strip(), val.strip().strip('"').strip("'")
            if key:
                os.environ[key] = val


def find_first_existing(paths: List[str]) -> str:
    for p in paths:
        if os.path.exists(p):
            return p
    raise FileNotFoundError(f"None of these files exist: {paths}")


def read_matching_pickle(path: str):
    """
    Load checkpoints / caches created with ``pd.read_pickle``.

    Conda ``base`` often cannot unpickle files saved under NumPy 2.x (e.g.
    ``No module named 'numpy._core.numeric'``). Use the **same interpreter**
    as the run that wrote the pickle, or upgrade NumPy inside conda.
    """
    try:
        return pd.read_pickle(path)
    except ModuleNotFoundError as e:
        es = str(e).lower()
        if "numpy" in es or "_core" in es:
            py = sys.executable or "python3"
            raise RuntimeError(
                f"Failed to load pickle {path!r}: {e}\n\n"
                "This usually means conda/base NumPy differs from the environment that saved the checkpoint.\n\n"
                "Fix — pick one:\n"
                f"  • Resume with Homebrew/system Python (often works):   python3 business_grade_matching.py --resume-from-step8\n"
                f"  • Or upgrade NumPy inside this env:                     conda install -y \"numpy>=2\" pandas\n"
                f"  • Or use the same interpreter that ran the pipeline:    {py} business_grade_matching.py --resume-from-step8\n\n"
                "Prefer `python3` (not conda `python`) if you mixed installs."
            ) from e
        raise


def safe_str(v) -> str:
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return ""
    return str(v).strip()


def sanitize_for_openpyxl(val):
    """Remove ASCII controls disallowed by openpyxl/Excel (keep TAB/LF/CR). GPT text can rarely inject e.g. 0x1A."""
    if not isinstance(val, str):
        return val
    return "".join(ch for ch in val if ord(ch) >= 32 or ch in "\t\n\r")


def sanitize_object_columns(frame: pd.DataFrame) -> pd.DataFrame:
    out = frame.copy()
    for col in out.columns:
        if out[col].dtype != object:
            continue
        out[col] = out[col].apply(
            lambda v: sanitize_for_openpyxl(v) if isinstance(v, str) else v
        )
    return out


def preprocess(text) -> str:
    """Step 1: clean, lowercase, normalize whitespace."""
    s = safe_str(text).lower()
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def safe_cos(v1, v2) -> float:
    try:
        return max(0.0, float(cosine_similarity([v1], [v2])[0][0]))
    except Exception:
        return 0.0


def to_list(v) -> List[str]:
    if v is None:
        return []
    if isinstance(v, list):
        return [safe_str(x) for x in v if safe_str(x)]
    return [safe_str(v)] if safe_str(v) else []


def parse_json(text: str):
    t = (text or "").strip()
    if t.startswith("```"):
        t = re.sub(r"^```(?:json)?", "", t).strip()
        t = re.sub(r"```$", "", t).strip()
    try:
        return json.loads(t)
    except Exception:
        return None


def call_json(client: OpenAI, model: str, system: str, payload: dict) -> dict:
    rsp = client.chat.completions.create(
        model=model,
        temperature=0,
        response_format={"type": "json_object"},
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": json.dumps(payload)},
        ],
    )
    parsed = parse_json(rsp.choices[0].message.content or "")
    if not parsed:
        raise RuntimeError("GPT returned non-JSON response")
    return parsed


def embed_batch(client: OpenAI, texts: List[str], model: str) -> List[List[float]]:
    """Embedding API with parallel chunk requests (much faster than one-at-a-time)."""
    bs = 128

    def one_slice(lo: int) -> Tuple[int, List[List[float]]]:
        chunk = texts[lo: lo + bs]
        inp = [t if t.strip() else "n/a" for t in chunk]
        r = client.embeddings.create(model=model, input=inp)
        return lo, [d.embedding for d in r.data]

    lows = list(range(0, len(texts), bs))
    if len(lows) == 1 or PARALLEL_EMBED_CHUNKS < 2:
        _, embs = one_slice(lows[0])
        all_embs = embs[:]
        for lo in lows[1:]:
            _, part = one_slice(lo)
            all_embs.extend(part)
        return all_embs

    buckets: Dict[int, List[List[float]]] = {}
    workers = min(PARALLEL_EMBED_CHUNKS, len(lows))
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futs = [ex.submit(one_slice, lo) for lo in lows]
        for fut in tqdm(
            as_completed(futs),
            desc="  Embedding chunks",
            total=len(lows),
            leave=False,
        ):
            lo, chunk_embs = fut.result()
            buckets[lo] = chunk_embs
    merged: List[List[float]] = []
    for lo in sorted(buckets.keys()):
        merged.extend(buckets[lo])
    return merged


def cosine_similarity_matrix(comp_emb: np.ndarray, opp_emb: np.ndarray) -> np.ndarray:
    """Cosine similarities (n_comp × n_opp), clipped like safe_cos. NaN-safe for zero norms."""
    comp_emb = np.asarray(comp_emb, dtype=np.float64)
    opp_emb = np.asarray(opp_emb, dtype=np.float64)
    eps = np.float64(1e-12)
    cn = np.linalg.norm(comp_emb, axis=1, keepdims=True)
    on = np.linalg.norm(opp_emb, axis=1, keepdims=True)
    c = np.divide(comp_emb, cn, out=np.zeros_like(comp_emb, dtype=np.float64), where=(cn >= eps))
    o = np.divide(opp_emb, on, out=np.zeros_like(opp_emb, dtype=np.float64), where=(on >= eps))
    raw = np.dot(c, o.T)
    raw = np.nan_to_num(raw, nan=0.0, posinf=0.0, neginf=0.0)
    return np.clip(raw, 0.0, 1.0)


def find_id_column(df: pd.DataFrame, hint: str) -> str:
    for cand in [hint, "id", "ID", "Id", f"{hint}_id", f"{hint}Id"]:
        if cand in df.columns:
            return cand
    return ""


def _normalize_header_key(name: object) -> str:
    """Lowercase + collapse whitespace; strip Excel noise from headers."""
    s = " ".join(str(name).strip().split())
    return s.casefold()


def normalize_company_cols(df: pd.DataFrame) -> pd.DataFrame:
    """
    Map KPMG / Excel variants to canonical names. Headers are stripped first;
    matching is **case-insensitive** so ``Company Name`` / ``COMPANY NAME`` /
    `` Company name `` all become ``company_name``.
    """
    df = df.copy()
    # Excel often injects NBSP (U+00A0) and zero-width space into “Company Name” headers.
    df.columns = [
        str(c).replace("\xa0", " ").replace("\u200b", "").strip()
        for c in df.columns
    ]

    # Normalized key → canonical column name used in the pipeline
    aliases = {
        "company name": "company_name",
        "companyname": "company_name",
        "name of company": "company_name",
        "organisation name": "company_name",
        "organization name": "company_name",
        "company profile": "company_profile",
        "product/services": "product_services",
        "product and services": "product_services",
        "product & services": "product_services",
        "products/services": "product_services",
        "products and services": "product_services",
        "product_services": "product_services",
        "product / services": "product_services",
        "sector": "Sector",
        "sectors": "Sector",
        "industry": "industry_field",
        "industries": "industry_field",
        "most active business unit": "business_unit_field",
        "business unit": "business_unit_field",
        "headquarters": "hq_field",
        "hq": "hq_field",
        "country": "country_field",
        "country of headquarters": "country_field",
        "website": "website_field",
        "website url": "website_field",
        "web site": "website_field",
    }

    new_cols: List[str] = []
    for c in df.columns:
        nk = _normalize_header_key(c)
        new_cols.append(aliases.get(nk, c))
    df.columns = new_cols

    missing = [c for c in ["company_name", "company_profile", "product_services", "Sector"] if c not in df.columns]
    if missing:
        raise KeyError(f"Company data missing columns: {missing}")
    return df


def normalize_opp_cols(df: pd.DataFrame) -> pd.DataFrame:
    required = [
        "What is the opportunity name?",
        "What is the opportunity description?",
        "Sector",
    ]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise KeyError(f"Opportunity data missing columns: {missing}")
    return df


def normalize_opportunities_for_pipeline(raw: pd.DataFrame) -> pd.DataFrame:
    """
    Prefer ``opportunities_structured.xlsx`` snake_case sheet; otherwise legacy questions.
    JSON list columns stored as strings in Excel: required_capabilities, etc.
    """
    df = raw.copy()
    df.columns = [str(c).replace("\xa0", " ").replace("\u200b", "").strip() for c in df.columns]
    structured = "opportunity_name" in df.columns and "required_capabilities" in df.columns
    json_cols = [
        "required_capabilities",
        "required_inputs",
        "adjacent_value_chain_sectors",
        "precedent_industries",
        "capability_keywords",
    ]
    if structured:
        df["What is the opportunity name?"] = df["opportunity_name"].map(safe_str)
        df["Sector"] = df["sector"].map(safe_str)
        df["What is the opportunity description?"] = df["opportunity_description"].map(safe_str)
        df["What are the investment highlights?"] = df["investment_highlights"].fillna("").map(safe_str)
        df["What is the value proposition of this opportunity?"] = (
            df["value_proposition"].fillna("").map(safe_str)
        )
        df["What are the key demand drivers?"] = df["demand_drivers"].fillna("").map(safe_str)
        df["Who are the key players in this sector or project?"] = df["key_players"].fillna("").map(safe_str)
        df["What materials are involved or required in the project?"] = (
            df["materials_required"].fillna("").map(safe_str)
        )
        df["Market data"] = df["market_data"].fillna("").map(safe_str)
        df["Cost structure"] = df["cost_structure"].fillna("").map(safe_str)
        df["Government incentives"] = df["government_incentives"].fillna("").map(safe_str)
        df["Risks and mitigations"] = df["risks_and_mitigations"].fillna("").map(safe_str)
        df["Investment locations"] = df["investment_locations"].fillna("").map(safe_str)
        normalize_opp_cols(df)
        return df

    df = normalize_opp_cols(df)
    for jc in json_cols:
        if jc not in df.columns:
            df[jc] = "[]"
    return df


def concat_matching_excel_sheets(
    path: str,
    normalize,
    role: str,
) -> pd.DataFrame:
    """
    Read **every worksheet** in a workbook (``sheet_name=None``) and stack only
    those that become valid after ``normalize`` (same required columns).

    ``pd.read_excel(path)`` alone only loads the first sheet — if KPMG data splits
    companies across tabs, you would silently see e.g. 200 rows instead of 3,200+.
    """
    book = pd.read_excel(path, sheet_name=None)

    chunks: List[pd.DataFrame] = []
    used: List[str] = []
    for sheet_name, raw in book.items():
        if raw is None or getattr(raw, "empty", True):
            continue
        try:
            chunks.append(normalize(raw))
            used.append(str(sheet_name))
        except KeyError as e:
            cols = [str(c).strip() for c in raw.columns[:25]]
            print(
                f"  [{role}] skipped tab {sheet_name!r} (missing/renamed columns): {e}; "
                f"first columns: {cols!r}",
            )
            continue

    if not chunks:
        raise ValueError(
            f"{role}: no tab in {path!r} has the required columns "
            f"(after header rename). Expected columns depend on {role} schema.",
        )

    out = pd.concat(chunks, ignore_index=True)
    print(f"  [{role}] Excel tabs merged: {used} → {len(out)} rows total")
    return out


def canonicalize_sector(s: str) -> str:
    """Cheap canonicalization to handle 'Healthcare and Life Sciences' vs 'Healthcare & Life Sciences'."""
    x = preprocess(s)
    x = x.replace(" and ", " ").replace(" & ", " ")
    return x.strip()


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — SECTOR ONTOLOGY EXPANSION
# ─────────────────────────────────────────────────────────────────────────────
ONTOLOGY_SYSTEM = """\
You are a sector taxonomy expert. For each input sector, produce a small
set of related sub-domains and adjacent domains that share value-chain or
capability overlap. Examples:

  CleanTech         → WaterTech, Desalination, Renewable Energy, Recycling
  ICT               → ICT Hardware, Telecommunications, 5G Equipment, Networking
  Industrial Mfg    → Heavy Industry, Steel, Petrochemicals, Manufacturing Equipment
  Healthcare        → Life Sciences, MedTech, Pharma, Diagnostics, Hospital Services

Return strict JSON only.
"""

ONTOLOGY_SCHEMA = {
    "results": [{
        "sector": "string",
        "canonical": "string  // one short canonical name",
        "synonyms": ["string  // alternate spellings, abbreviations"],
        "sub_domains": ["string  // 3–6 named sub-sectors"],
        "adjacent_domains": ["string  // 3–6 adjacent sectors w/ value-chain crossover"],
    }]
}


def load_ontology_cache(path: str) -> Dict[str, dict]:
    if os.path.exists(path):
        try:
            with open(path) as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_ontology_cache(path: str, cache: Dict[str, dict]) -> None:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w") as f:
        json.dump(cache, f, indent=2)


def expand_sectors_via_gpt(client: OpenAI, sectors: List[str]) -> Dict[str, dict]:
    """Batched GPT call. Returns {sector: {canonical, synonyms, sub_domains, adjacent_domains}}."""
    if not sectors:
        return {}
    payload = {
        "sectors": sectors,
        "instructions": "Produce expansion for each sector. Return strict JSON.",
        "output_schema": ONTOLOGY_SCHEMA,
    }
    parsed = call_json(client, ONTOLOGY_MODEL, ONTOLOGY_SYSTEM, payload)
    out: Dict[str, dict] = {}
    for r in parsed.get("results", []):
        s = safe_str(r.get("sector"))
        if not s:
            continue
        out[s] = {
            "canonical":        safe_str(r.get("canonical")) or s,
            "synonyms":         to_list(r.get("synonyms")),
            "sub_domains":      to_list(r.get("sub_domains")),
            "adjacent_domains": to_list(r.get("adjacent_domains")),
        }
    return out


def build_ontology(client: OpenAI, sectors: List[str], cache_path: str) -> Dict[str, Set[str]]:
    """
    For each sector, return a set of canonicalized tokens describing its
    semantic neighbourhood (the sector itself, its canonical name, synonyms,
    sub-domains and adjacent domains).
    """
    cache = load_ontology_cache(cache_path)
    unique = sorted({s for s in sectors if safe_str(s)})
    todo = [s for s in unique if s not in cache]

    if todo:
        # Batch in groups of 24 to keep prompts small
        for i in tqdm(range(0, len(todo), 24), desc="  GPT ontology expansion"):
            batch = todo[i: i + 24]
            try:
                got = expand_sectors_via_gpt(client, batch)
            except Exception as e:
                print(f"  Ontology batch failed: {e}")
                got = {}
            for s in batch:
                cache[s] = got.get(s, {
                    "canonical": s,
                    "synonyms": [],
                    "sub_domains": [],
                    "adjacent_domains": [],
                })
        save_ontology_cache(cache_path, cache)

    expanded: Dict[str, Set[str]] = {}
    for s in unique:
        info = cache.get(s, {})
        bag = {canonicalize_sector(s)}
        bag.add(canonicalize_sector(info.get("canonical", s)))
        for tok in info.get("synonyms", []) + info.get("sub_domains", []) + info.get("adjacent_domains", []):
            bag.add(canonicalize_sector(tok))
        bag.discard("")
        expanded[s] = bag
    return expanded


def sectors_overlap(a: str, b: str, expanded: Dict[str, Set[str]]) -> bool:
    """Step 3: pair passes filter if expanded sector sets overlap."""
    ea = expanded.get(a, {canonicalize_sector(a)})
    eb = expanded.get(b, {canonicalize_sector(b)})
    return bool(ea & eb)


# ─────────────────────────────────────────────────────────────────────────────
# ENTITY EXTRACTION (used to build product/service embedding text — Step 5)
# ─────────────────────────────────────────────────────────────────────────────
OPP_EXTRACT_SCHEMA = {
    "results": [{
        "id": "string",
        "required_products": ["string"],
        "required_services": ["string"],
        "required_materials": ["string"],
        "required_capabilities": ["string"],
    }]
}

COMP_EXTRACT_SCHEMA = {
    "results": [{
        "id": "string",
        "actual_products": ["string"],
        "actual_services": ["string"],
        "capabilities": ["string"],
    }]
}


def extract_batch(client: OpenAI, items: List[dict], kind: str) -> Dict[str, dict]:
    schema = OPP_EXTRACT_SCHEMA if kind == "opportunity" else COMP_EXTRACT_SCHEMA
    instruction = (
        "Extract specific, concrete entities — products, materials, capabilities. "
        "Avoid generic terms like 'solutions', 'services', 'management'. Return strict JSON."
    )
    payload = {"items": items, "instructions": instruction, "output_schema": schema}
    system  = "You extract structured business entities from industrial descriptions."
    parsed  = call_json(client, EXTRACTION_MODEL, system, payload)
    return {
        str(r.get("id", "")).strip(): r
        for r in parsed.get("results", [])
        if str(r.get("id", "")).strip()
    }


def opp_entity_text(raw: dict) -> str:
    parts = (
        to_list(raw.get("required_products"))
        + to_list(raw.get("required_services"))
        + to_list(raw.get("required_materials"))
        + to_list(raw.get("required_capabilities"))
    )
    return " ".join(parts) or "n/a"


def comp_entity_text(raw: dict) -> str:
    parts = (
        to_list(raw.get("actual_products"))
        + to_list(raw.get("actual_services"))
        + to_list(raw.get("capabilities"))
    )
    return " ".join(parts) or "n/a"


def load_entity_extraction_cache() -> Dict[str, Dict[str, dict]]:
    if os.path.isfile(ENTITY_EXTRACTION_CACHE):
        try:
            blob = read_matching_pickle(ENTITY_EXTRACTION_CACHE)
            return {
                "company":      dict(blob.get("company", {}) or {}),
                "opportunity":  dict(blob.get("opportunity", {}) or {}),
            }
        except Exception:
            pass
    return {"company": {}, "opportunity": {}}


def save_entity_extraction_cache(blob: Dict[str, Dict[str, dict]]) -> None:
    os.makedirs(os.path.dirname(ENTITY_EXTRACTION_CACHE) or ".", exist_ok=True)
    pd.to_pickle(blob, ENTITY_EXTRACTION_CACHE)


def populate_entity_texts_cached_parallel(
    client: OpenAI,
    items: List[dict],
    kind: str,
    cache_bucket: Dict[str, dict],
    raw_to_text,
) -> Dict[str, str]:
    """
    Fill ``cache_bucket`` with GPT extraction results where missing,
    optionally in parallel batches. Returns {_key → embedding text}.
    """
    missing_items: List[dict] = []
    for it in items:
        cid = str(it["id"]).strip()
        if cid not in cache_bucket:
            missing_items.append(it)

    batches: List[List[dict]] = []
    for i in range(0, len(missing_items), EXTRACT_BATCH):
        batches.append(missing_items[i: i + EXTRACT_BATCH])

    if batches:
        if len(batches) == 1 or PARALLEL_EXTRACT_BATCH < 2:
            for batch in tqdm(batches, desc=f"  {kind[:3]} extraction", leave=False):
                try:
                    resp = extract_batch(client, batch, kind)
                except Exception as e:
                    print(f"    extraction failed {kind}: {e}")
                    resp = {}
                for b in batch:
                    cache_bucket.setdefault(b["id"], resp.get(b["id"], {}))
        else:
            workers = min(PARALLEL_EXTRACT_BATCH, len(batches))

            def run_one(batch: List[dict]) -> None:
                try:
                    resp = extract_batch(client, batch, kind)
                except Exception as e:
                    print(f"    extraction failed {kind} batch: {e}")
                    resp = {}
                for b in batch:
                    cache_bucket.setdefault(b["id"], resp.get(b["id"], {}))

            with ThreadPoolExecutor(max_workers=workers) as ex:
                fut_list = [ex.submit(run_one, batch) for batch in batches]
                for fut in tqdm(
                    as_completed(fut_list),
                    total=len(batches),
                    desc=f"  {kind[:3]} batches",
                    leave=False,
                ):
                    fut.result()

    out_txt: Dict[str, str] = {}
    for it in items:
        cid = it["id"]
        out_txt[cid] = raw_to_text(cache_bucket.get(cid, {}))
    return out_txt


# ─────────────────────────────────────────────────────────────────────────────
# STEP 6 — GPT-BASED VALIDATION (analyst-grade narrative)
# ─────────────────────────────────────────────────────────────────────────────
DECISION_SYSTEM_PROMPT = """\
You are a senior deal-sourcing analyst evaluating whether a company can
realistically fulfill an industrial or infrastructure opportunity. Your output
is read by a business-development team that needs to (a) decide pursue/not
pursue and (b) know exactly what to do next if they pursue.

CORE QUESTION
  "Can this company realistically fulfill this opportunity — directly or as a
  cross-sector adjacency with proven crossover value?"

TONE — investment-memo style
  Sharp. Specific. Cite real, named evidence — actual products, divisions,
  certifications, geographies, customer references. Never use generic phrasing
  like "well-positioned", "comprehensive solutions", or "industry leader"
  without naming the specific capability behind the claim.

  When the opportunity description mentions a regional context (e.g. KSA,
  Saudi Vision 2030, MENA, etc.), incorporate that context into the narrative.
  When it mentions specific demand drivers, name them.

DECISION (binary)
  ai_decision = "Yes"  → realistically can fulfill the opportunity, directly or
                          via a defensible cross-sector adjacency. Has named
                          capability that serves the requirement, or close
                          adjacency with proven crossover value.
  ai_decision = "No"   → not viable. Sector or capability mismatch with no
                          realistic participation route. Generic word overlap
                          is not a fit.

You DO NOT default to rejection. About half of valid matches in this domain
come from cross-sector adjacencies with strong product overlap. Articulate
the path when it exists.

NARRATIVE FIELDS — these are the deliverable, not stub fragments.

ai_explanation     2–4 sentences, flowing analyst paragraph in the style of an
                   investment memo. State (a) what the company is and what it
                   does, citing specific products, divisions, history if
                   known; (b) why it is or is not relevant to this specific
                   opportunity, citing the demand driver or capability gap;
                   (c) the strategic implication — technology transfer,
                   regional market entry, capability extension, etc. For
                   "No" decisions, lead with "No." then explain.

ai_insight         One non-obvious strategic observation — an adjacency that
                   unlocks the opportunity, a hidden risk, a partnership angle,
                   a regulatory consideration, or a capability transfer angle.

suggested_plan     Exactly 3 concrete, sequenced engagement steps. Each step
                   is a specific action a BD team can take this quarter. Name
                   tools, products, or activities. No platitudes — no "leverage
                   synergies", "explore opportunities", or "build relationships".

match_reason       Exactly 3 specific, evidence-cited reasons. Each names a
                   real capability, market fact, or product. For "No"
                   decisions, give 3 reasons why this is a non-fit.

EVERY narrative field is required for both Yes and No decisions.

OUTPUT
Return one JSON object per candidate, wrapped in {"results": [...]}. Each
object has exactly these fields:

{
  "candidate_id": "string",
  "ai_decision": "Yes" | "No",
  "ai_explanation": "2–4 sentence analyst paragraph",
  "ai_insight": "one non-obvious strategic observation",
  "suggested_plan": ["step 1", "step 2", "step 3"],
  "match_reason": ["reason 1", "reason 2", "reason 3"]
}
"""


def gpt_decide_batch(
    client: OpenAI,
    opp_context: dict,
    candidates: List[dict],
) -> Dict[str, dict]:
    payload = {
        "opportunity": opp_context,
        "candidates": candidates,
        "instruction": (
            "For EACH candidate, produce the full JSON object defined in the system prompt. "
            "Return exactly: {\"results\": [...]}. Every narrative field must be specific "
            "and named — investment-memo tone, no generic filler."
        ),
    }
    parsed = call_json(client, DECISION_MODEL, DECISION_SYSTEM_PROMPT, payload)
    return {
        str(r.get("candidate_id", "")).strip(): r
        for r in parsed.get("results", [])
        if str(r.get("candidate_id", "")).strip()
    }


def safe_decision(r: dict, fallback_cid: str) -> dict:
    """Normalize LLM output — only Yes | No allowed (Filtered is pipeline-only)."""
    raw_d = str(r.get("ai_decision", "No")).strip().casefold()
    if raw_d == "yes":
        decision = "Yes"
    elif raw_d == "no":
        decision = "No"
    else:
        decision = "No"

    plan = to_list(r.get("suggested_plan"))[:3]
    while len(plan) < 3:
        plan.append("")

    reasons = to_list(r.get("match_reason"))[:3]
    while len(reasons) < 3:
        reasons.append("")

    return {
        "candidate_id":    str(r.get("candidate_id", fallback_cid)).strip(),
        "ai_decision":     decision,
        "ai_explanation":  safe_str(r.get("ai_explanation")),
        "ai_insight":      safe_str(r.get("ai_insight")),
        "suggested_plan":  plan,
        "match_reason":    reasons,
    }


def _stub_match_reasons_filtered() -> List[str]:
    return [
        "Filtered by multi-signal triage — no qualifying routing signal exceeded thresholds.",
        f"(Capability keywords≥{TRIAGE_MIN_CAPABILITY_KEYWORD_HITS}, inputs≥{TRIAGE_MIN_INPUT_OVERLAP_HITS}, "
        f"signal-product≥{TRIAGE_MIN_SIGNAL_PRODUCT_SIM}, signal-profile≥{TRIAGE_MIN_SIGNAL_PROFILE_SIM}.)",
        "See column triage_reason for this row's signal values.",
    ]


def stamp_sector_gate_filtered(df: pd.DataFrame) -> None:
    """Pairs that never ran the LLM (multi-signal triage): ``ai_decision`` = Filtered."""
    if "_send_to_gpt" not in df.columns:
        return
    fmask = ~df["_send_to_gpt"].fillna(False).astype(bool)
    if not fmask.any():
        return
    df.loc[fmask, "ai_decision"] = "Filtered"
    df.loc[fmask, "ai_score"] = 0
    df.loc[fmask, "ai_explanation"] = (
        "Filtered by multi-signal triage — company–opportunity pair did not satisfy any routing signal "
        f"(ontology sector overlap, adjacent value-chain sector, capability keywords≥"
        f"{TRIAGE_MIN_CAPABILITY_KEYWORD_HITS}, required-input overlap≥{TRIAGE_MIN_INPUT_OVERLAP_HITS}, "
        f"structured product embedding≥{TRIAGE_MIN_SIGNAL_PRODUCT_SIM}, or profile–description embedding≥"
        f"{TRIAGE_MIN_SIGNAL_PROFILE_SIM}). See triage_reason."
    )
    df.loc[fmask, "ai_insight"] = ""
    df.loc[fmask, "suggested_plan"] = df.loc[fmask, "suggested_plan"].apply(lambda _: [])
    df.loc[fmask, "match_reason"] = df.loc[fmask, "match_reason"].apply(
        lambda _: _stub_match_reasons_filtered(),
    )


def write_output_and_summary(out: pd.DataFrame, xlsx_path: str = OUTPUT_FILE) -> None:
    """Sanitize strings, write xlsx, optional formatting, print summary."""
    os.makedirs(os.path.dirname(xlsx_path) or ".", exist_ok=True)
    safe = sanitize_object_columns(out)
    safe.to_excel(xlsx_path, index=False)

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        wb = load_workbook(xlsx_path)
        ws = wb.active
        headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        narrative = {
            "ai_explanation",
            "ai_insight",
            "suggested_plan",
            "match_reason",
            "company_name",
            "opportunity_name",
            "company_profile",
            "opportunity_description",
            "sector_inference_evidence",
            "triage_reason",
        }
        widths = {
            "id": 6,
            "companyId": 10,
            "opportunityId": 12,
            "company_name": 26,
            "opportunity_name": 38,
            "company_profile": 56,
            "opportunity_description": 56,
            "company_sector": 18,
            "opportunity_sector": 18,
            "sector_similarity": 12,
            "ontology_sector_overlap": 22,
            "profile_similarity": 10,
            "product_similarity": 10,
            "ai_score": 8,
            "ai_decision": 12,
            "final_score": 12,
            "ai_explanation": 70,
            "rank": 6,
            "ai_insight": 60,
            "suggested_plan": 60,
            "match_reason": 60,
            "sector_was_inferred": 14,
            "sector_inference_source": 16,
            "sector_inference_evidence": 52,
            "signal_sector_overlap": 12,
            "signal_value_chain_adjacent": 14,
            "signal_capability_keyword_count": 16,
            "signal_input_overlap_count": 14,
            "signal_product_similarity": 14,
            "signal_profile_similarity": 14,
            "triage_pass": 10,
            "triage_reason": 56,
        }
        for c, name in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(c)].width = widths.get(name, 16)
            if name in narrative:
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        if "ai_decision" in headers:
            dec_col = headers.index("ai_decision") + 1
            yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            no_fill  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            filtered_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=dec_col).value
                if v == "Yes":
                    ws.cell(row=r, column=dec_col).fill = yes_fill
                elif v == "No":
                    ws.cell(row=r, column=dec_col).fill = no_fill
                elif v == "Filtered":
                    ws.cell(row=r, column=dec_col).fill = filtered_fill
        ws.freeze_panes = "A2"
        wb.save(xlsx_path)
    except Exception as e:
        print(f"  (Excel formatting skipped: {e})")

    print(f"\nOutput saved : {xlsx_path}")
    print(f"Total rows   : {len(safe)}")
    print(f"\nai_decision breakdown:")
    for d, cnt in safe["ai_decision"].value_counts().items():
        print(f"  {d:<5} {cnt}")

    if "ontology_sector_overlap" in safe.columns:
        yes_in = (
            ((safe["ai_decision"] == "Yes") & (safe["ontology_sector_overlap"].astype(int) == 1))
        ).sum()
        yes_cross = (
            ((safe["ai_decision"] == "Yes") & (safe["ontology_sector_overlap"].astype(int) == 0))
        ).sum()
        print(f"\nYes matches: {yes_in} with ontology sector-overlap signal + "
              f"{yes_cross} without overlap signal (routing via multi-signal triage; see triage_reason)")
    else:
        print("\nYes matches summary: ontology_overlap column absent (legacy export).")

    print(f"\nfinal_score percentiles:")
    for p in [0.10, 0.25, 0.50, 0.75, 0.90, 0.99]:
        print(f"  p{int(p*100):>2}: {safe['final_score'].quantile(p):.3f}")


OPP_NAME_COL = "What is the opportunity name?"
OPP_DESC_COL = "What is the opportunity description?"


EXPORT_TEXT_COLS = (
    "company_name",
    "opportunity_name",
    "company_profile",
    "opportunity_description",
)


def enrich_pair_table_with_source_texts(df: pd.DataFrame) -> pd.DataFrame:
    """
    Pair rows normally carry ``company_name``, ``company_profile``, opportunity
    name/description from the Cartesian build. Older ``post_step6`` checkpoints
    may lack them — merge from the same workbooks using ``_comp_key`` /
    ``_opp_key``.

    We only skip merging when **all four** columns are already present; a single
    stray ``opportunity_name`` column must not short-circuit the rest.
    """
    df = df.copy()
    if all(c in df.columns for c in EXPORT_TEXT_COLS):
        return df
    if "_comp_key" not in df.columns or "_opp_key" not in df.columns:
        print("  Note: checkpoint has no _comp_key/_opp_key; cannot attach source text columns.")
        return df
    try:
        company_file = find_first_existing(COMPANY_FILE_CANDIDATES)
        opp_file = find_first_existing(OPPORTUNITY_FILE_CANDIDATES)
    except FileNotFoundError as e:
        print(f"  Note: could not load workbooks to enrich rows ({e}).")
        return df

    companies_fb = concat_matching_excel_sheets(company_file, normalize_company_cols, "Companies")
    opportunities_fb = concat_matching_excel_sheets(
        opp_file, normalize_opportunities_for_pipeline, "Opportunities",
    )
    companies_fb = companies_fb.reset_index(drop=True)
    opportunities_fb = opportunities_fb.reset_index(drop=True)
    companies_fb["_key"] = companies_fb.index.map(lambda i: f"C{i}")
    opportunities_fb["_key"] = opportunities_fb.index.map(lambda i: f"O{i}")

    cside = companies_fb[["_key", "company_name", "company_profile"]].rename(
        columns={"_key": "_comp_key"},
    )
    oside = opportunities_fb[["_key", OPP_NAME_COL, OPP_DESC_COL]].rename(
        columns={"_key": "_opp_key", OPP_NAME_COL: "opportunity_name", OPP_DESC_COL: "opportunity_description"},
    )
    if not {"company_name", "company_profile"}.issubset(df.columns):
        df = df.merge(cside, on="_comp_key", how="left")
    if not {"opportunity_name", "opportunity_description"}.issubset(df.columns):
        df = df.merge(oside, on="_opp_key", how="left")
    for col in EXPORT_TEXT_COLS:
        if col in df.columns:
            df[col] = df[col].fillna("").map(lambda x: safe_str(x))
        else:
            df[col] = ""
    return df


def _verify_export_rank_semantics(frame: pd.DataFrame) -> None:
    """Sanity-check per-opportunity rank vs (final_score, product_similarity) ordering."""
    bad_mono = 0
    top_not_yes = 0
    for oid, sub in frame.groupby("opportunityId", sort=False):
        sub = sub.sort_values("rank")
        fc = pd.to_numeric(sub["final_score"], errors="coerce").fillna(0.0).astype(float).to_numpy()
        pc = pd.to_numeric(sub["product_similarity"], errors="coerce").fillna(0.0).astype(float).to_numpy()
        seq = [(fc[i], pc[i]) for i in range(len(fc))]
        if seq != sorted(seq, key=lambda t: (-t[0], -t[1])):
            bad_mono += 1
        if ((sub["ai_decision"].astype(str) == "Yes").any()
                and str(sub.iloc[0]["ai_decision"]).strip() != "Yes"):
            top_not_yes += 1
    n_opp = int(frame["opportunityId"].nunique())
    msg = (
        f"  Rank QA: opportunities with ordering ≠ (final_score↓, product↓): {bad_mono}/{n_opp}; "
        f"where rank=1 not Yes but opp has ≥1 Yes: {top_not_yes}/{n_opp}"
    )
    print(msg)


def _ensure_pair_sector_inference_defaults(df: pd.DataFrame) -> pd.DataFrame:
    """Older checkpoints may omit sector-inference audit columns."""
    out = df
    if "sector_was_inferred" not in out.columns:
        out["sector_was_inferred"] = False
    out["sector_was_inferred"] = out["sector_was_inferred"].fillna(False).astype(bool)
    if "sector_inference_source" not in out.columns:
        out["sector_inference_source"] = "none"
    out["sector_inference_source"] = out["sector_inference_source"].fillna("none").map(lambda x: safe_str(x))
    if "sector_inference_evidence" not in out.columns:
        out["sector_inference_evidence"] = ""
    out["sector_inference_evidence"] = out["sector_inference_evidence"].fillna("").map(lambda x: safe_str(x))
    return out


def _ensure_pair_triage_defaults(df: pd.DataFrame) -> pd.DataFrame:
    if "signal_sector_overlap" not in df.columns:
        df["signal_sector_overlap"] = 0
    df["signal_sector_overlap"] = pd.to_numeric(
        df["signal_sector_overlap"], errors="coerce",
    ).fillna(0).astype(int)
    if "signal_value_chain_adjacent" not in df.columns:
        df["signal_value_chain_adjacent"] = False
    df["signal_value_chain_adjacent"] = df["signal_value_chain_adjacent"].fillna(False).astype(bool)
    if "signal_capability_keyword_count" not in df.columns:
        df["signal_capability_keyword_count"] = 0
    df["signal_capability_keyword_count"] = pd.to_numeric(
        df["signal_capability_keyword_count"], errors="coerce",
    ).fillna(0).astype(int)
    if "signal_input_overlap_count" not in df.columns:
        df["signal_input_overlap_count"] = 0
    df["signal_input_overlap_count"] = pd.to_numeric(
        df["signal_input_overlap_count"], errors="coerce",
    ).fillna(0).astype(int)
    if "signal_product_similarity" not in df.columns:
        df["signal_product_similarity"] = 0.0
    df["signal_product_similarity"] = pd.to_numeric(
        df["signal_product_similarity"], errors="coerce",
    ).fillna(0.0).astype(float)
    if "signal_profile_similarity" not in df.columns:
        df["signal_profile_similarity"] = 0.0
    df["signal_profile_similarity"] = pd.to_numeric(
        df["signal_profile_similarity"], errors="coerce",
    ).fillna(0.0).astype(float)
    if "triage_pass" not in df.columns:
        df["triage_pass"] = False
    df["triage_pass"] = df["triage_pass"].fillna(False).astype(bool)
    if "triage_reason" not in df.columns:
        df["triage_reason"] = ""
    df["triage_reason"] = df["triage_reason"].fillna("").map(lambda x: safe_str(x))
    return df


def print_sector_inference_reconciliation(
    out: pd.DataFrame,
    companies_summary: Optional[pd.DataFrame] = None,
) -> None:
    if "sector_was_inferred" not in out.columns:
        return
    by_c = out.groupby("companyId", sort=False).first(numeric_only=False)
    inf_m = by_c["sector_was_inferred"].fillna(False).astype(bool)
    n_inf = int(inf_m.sum())
    print("\n[Sector inference reconciliation]")
    print(f"  Companies (distinct companyId) with inferred flag: {n_inf}")
    if n_inf:
        sub = by_c.loc[inf_m]
        print("  sector_inference_source (company-level):")
        for k, v in sub["sector_inference_source"].fillna("none").value_counts().items():
            print(f"    {k}: {int(v)}")
    if companies_summary is not None and all(
        c in companies_summary.columns for c in ("Sector", "_company_id", "sector_was_inferred")
    ):
        c = companies_summary
        blank = c["Sector"].astype(str).map(safe_str).eq("")
        n_blank = int(blank.sum())
        inf_comp = c["sector_was_inferred"].fillna(False).astype(bool)
        print(f"  Workbook rows with blank Sector: {n_blank}; inferred rows: {int((blank & inf_comp).sum())}")
        inf_ids = set(c.loc[inf_comp, "_company_id"].astype(int).tolist())
        yes_ids = set(out.loc[out["ai_decision"].eq("Yes"), "companyId"].astype(int).unique().tolist())
        print(f"  Inferred-sector companies with ≥1 Yes pair: {len(inf_ids & yes_ids)}")
        filt_only = 0
        for cid, grp in out.groupby("companyId", sort=False):
            if int(cid) not in inf_ids:
                continue
            if grp["ai_decision"].eq("Filtered").all():
                filt_only += 1
        print(f"  Inferred-sector companies where all pairs are Filtered: {filt_only}")


def finalize_ranking_and_export(
    df: pd.DataFrame,
    companies_summary: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """Step 8 + build output sheet + pre-export pickle + Excel (strict pipeline tail)."""
    # ── STEP 8 — Ranking and Scoring ──────────────────────────────────────────
    print("[Step 8/8] Final score and ranking...")
    df = df.copy()
    df = enrich_pair_table_with_source_texts(df)
    for col in EXPORT_TEXT_COLS:
        if col not in df.columns:
            df[col] = ""
    df = _ensure_pair_sector_inference_defaults(df)
    df = _ensure_pair_triage_defaults(df)
    stamp_sector_gate_filtered(df)

    prof = pd.to_numeric(df["profile_similarity"], errors="coerce").fillna(0.0).astype(np.float64)
    prod = pd.to_numeric(df["product_similarity"], errors="coerce").fillna(0.0).astype(np.float64)
    df["sector_similarity"] = pd.concat([prof, prod], axis=1).max(axis=1).astype(float).round(3)
    if "_sector_ontology_overlap" not in df.columns:
        if "_pass_reason" in df.columns:
            df["_sector_ontology_overlap"] = df["_pass_reason"].eq("sector_overlap").astype(int)
        else:
            df["_sector_ontology_overlap"] = 0
    df["_sector_ontology_overlap"] = pd.to_numeric(
        df["_sector_ontology_overlap"], errors="coerce",
    ).fillna(0).astype(int)

    df["final_score"] = (
        W_AI * df["ai_score"].astype(float)
        + W_PROFILE * df["profile_similarity"].astype(float)
        + W_PRODUCT * df["product_similarity"].astype(float)
    ).round(3)

    # ROW_NUMBER() OVER (PARTITION BY opportunityId ORDER BY final_score DESC,
    #   product_similarity DESC) — deterministic last key via original row order.
    df["_prod_rank_tiebreak"] = (
        pd.to_numeric(df["product_similarity"], errors="coerce").fillna(0.0).astype(np.float64)
    )
    df["_rank_stable"] = np.arange(len(df), dtype=np.int64)
    df = df.sort_values(
        by=["opportunityId", "final_score", "_prod_rank_tiebreak", "_rank_stable"],
        ascending=[True, False, False, True],
    ).reset_index(drop=True)
    df = df.drop(columns=["_prod_rank_tiebreak", "_rank_stable"], errors="ignore")
    df["rank"] = df.groupby("opportunityId", sort=False).cumcount() + 1

    _verify_export_rank_semantics(df)

    df = df.sort_values(by=["companyId", "rank"]).reset_index(drop=True)
    df.insert(0, "id", range(1, len(df) + 1))

    df["suggested_plan_str"] = df["suggested_plan"].apply(lambda xs: json.dumps(xs, ensure_ascii=False))
    df["match_reason_str"] = df["match_reason"].apply(lambda xs: json.dumps(xs, ensure_ascii=False))

    out = pd.DataFrame({
        "id":                       df["id"].astype(int),
        "companyId":                df["companyId"].astype(int),
        "opportunityId":            df["opportunityId"].astype(int),
        "company_name":             df["company_name"].map(lambda x: safe_str(x)),
        "opportunity_name":         df["opportunity_name"].map(lambda x: safe_str(x)),
        "company_profile":          df["company_profile"].map(lambda x: safe_str(x)),
        "opportunity_description":  df["opportunity_description"].map(lambda x: safe_str(x)),
        "company_sector":           df["company_sector"],
        "opportunity_sector":        df["opportunity_sector"],
        "sector_similarity":         df["sector_similarity"].astype(float).round(3),
        "ontology_sector_overlap":   df["_sector_ontology_overlap"].astype(int),
        "profile_similarity":        df["profile_similarity"].astype(float).round(3),
        "product_similarity":        df["product_similarity"].astype(float).round(3),
        "ai_score":                 df["ai_score"].astype(int),
        "ai_decision":              df["ai_decision"],
        "final_score":              df["final_score"].astype(float).round(3),
        "ai_explanation":           df["ai_explanation"],
        "rank":                     df["rank"].astype(int),
        "ai_insight":               df["ai_insight"],
        "suggested_plan":           df["suggested_plan_str"],
        "match_reason":             df["match_reason_str"],
        "sector_was_inferred":      df["sector_was_inferred"].astype(bool),
        "sector_inference_source":  df["sector_inference_source"].map(lambda x: safe_str(x)),
        "sector_inference_evidence": df["sector_inference_evidence"].map(lambda x: safe_str(x)),
        "signal_sector_overlap":    df["signal_sector_overlap"].astype(int),
        "signal_value_chain_adjacent": df["signal_value_chain_adjacent"].astype(bool),
        "signal_capability_keyword_count": df["signal_capability_keyword_count"].astype(int),
        "signal_input_overlap_count": df["signal_input_overlap_count"].astype(int),
        "signal_product_similarity": df["signal_product_similarity"].astype(float).round(3),
        "signal_profile_similarity": df["signal_profile_similarity"].astype(float).round(3),
        "triage_pass":              df["triage_pass"].astype(bool),
        "triage_reason":            df["triage_reason"].map(lambda x: safe_str(x)),
    })

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    out.to_pickle(CHECKPOINT_PRE_EXPORT)
    print(f"  Pre-export checkpoint: {CHECKPOINT_PRE_EXPORT} (use --resume-export if .xlsx fails)")
    print(f"  Export columns ({len(out.columns)}): {', '.join(out.columns.astype(str).tolist())}")
    write_output_and_summary(out)
    print_sector_inference_reconciliation(out, companies_summary)
    return out


def resume_export_only() -> None:
    """Load pre-export checkpoint and write xlsx only (no API calls)."""
    if not os.path.isfile(CHECKPOINT_PRE_EXPORT):
        raise FileNotFoundError(
            f"No checkpoint at {CHECKPOINT_PRE_EXPORT!r}. Run the full pipeline first; "
            "the checkpoint is written just before Excel export."
        )
    print(f"Loading checkpoint: {CHECKPOINT_PRE_EXPORT}")
    out = read_matching_pickle(CHECKPOINT_PRE_EXPORT)
    stale_schema = (
        "ontology_sector_overlap" not in out.columns
        or "sector_similarity" not in out.columns
        or not pd.api.types.is_float_dtype(out["sector_similarity"].dtype)
    )
    if stale_schema and not os.path.isfile(CHECKPOINT_POST_STEP6):
        raise ValueError(
            f"{CHECKPOINT_PRE_EXPORT!r} needs regeneration (ontology_sector_overlap / "
            "float sector_similarity). Run: python3 business_grade_matching.py --resume-from-step8 "
            f"(requires {CHECKPOINT_POST_STEP6!r}), or rerun the full pipeline."
        )
    if (
        stale_schema or not set(EXPORT_TEXT_COLS).issubset(out.columns)
    ) and os.path.isfile(CHECKPOINT_POST_STEP6):
        print(
            "Pre-export pickle is stale (missing readable columns or new sector_similarity semantics). "
            f"Rebuilding from {CHECKPOINT_POST_STEP6!r} (Step 8 only, no API)…",
        )
        df6 = read_matching_pickle(CHECKPOINT_POST_STEP6)
        finalize_ranking_and_export(df6, companies_summary=None)
        return
    if not set(EXPORT_TEXT_COLS).issubset(out.columns):
        raise ValueError(
            f"Loaded {CHECKPOINT_PRE_EXPORT!r} is missing columns {EXPORT_TEXT_COLS}. "
            "Run: python3 business_grade_matching.py --resume-from-step8 "
            f"(needs {CHECKPOINT_POST_STEP6!r}) or re-run the full pipeline."
        )
    print(f"  Export columns ({len(out.columns)}): {', '.join(out.columns.astype(str).tolist())}")
    out = _ensure_pair_sector_inference_defaults(out)
    out = _ensure_pair_triage_defaults(out)
    write_output_and_summary(out)


def resume_from_step8_only() -> None:
    """Load post–Step 6 pair table; run Step 8 + pre-export pickle + xlsx (no API calls)."""
    if not os.path.isfile(CHECKPOINT_POST_STEP6):
        raise FileNotFoundError(
            f"No checkpoint at {CHECKPOINT_POST_STEP6!r}. Run the full pipeline through Step 6 first."
        )
    print(f"Loading checkpoint: {CHECKPOINT_POST_STEP6}")
    df = read_matching_pickle(CHECKPOINT_POST_STEP6)
    print("Resuming: Step 8 + export only (Steps 1–7 skipped).")
    finalize_ranking_and_export(df, companies_summary=None)


def _print_cli_help() -> None:
    print("Usage: python3 business_grade_matching.py [OPTION]")
    print("  (pickles from full runs align with NumPy 2.x; use python3 — not conda python — if unpickle fails)")
    print("  (default)                Full 8-step pipeline; requires OPENAI_API_KEY")
    print("  --dry-sector-stats       Load kpmgfile.xlsx only; print Sector / inference stats (no API).")
    print("  --resume-from-step8      After long GPT run: load Output/business_grade_post_step6.pkl,")
    print("                           run ranking + Excel only (no API).")
    print("  --resume-export, -r      If .xlsx fails: load Output/business_grade_pre_export.pkl,")
    print("                           sanitize + write workbook only (no API).")
    print("Related:")
    print("  extract_opportunities_structured.py  → build opportunities_structured.xlsx (Phase 1)")
    print("  calibration_report.py --baseline X --new Y   → Phase 4 regression / calibration")
    print("Company file must resolve to kpmgfile.xlsx (first match in COMPANY_FILE_CANDIDATES).")


def run_dry_sector_stats() -> None:
    """Load company workbook like the pipeline and print sector / inference diagnostics (no API)."""
    company_file = find_first_existing(COMPANY_FILE_CANDIDATES)
    if os.path.basename(company_file) != "kpmgfile.xlsx":
        raise ValueError(
            "Configured to use kpmgfile.xlsx as company source. Found "
            f"{company_file!r} instead. Add kpmgfile.xlsx to the project directory."
        )
    companies = concat_matching_excel_sheets(company_file, normalize_company_cols, "Companies")
    for oc in (
        "industry_field",
        "business_unit_field",
        "hq_field",
        "country_field",
        "website_field",
    ):
        if oc not in companies.columns:
            companies[oc] = ""

    sector_nonempty = companies["Sector"].astype(str).map(safe_str).ne("")
    n_blank = int((~sector_nonempty).sum())
    vocab = closed_sector_vocabulary(companies)
    try:
        with open(SECTOR_INFERENCE_CACHE_FILE, encoding="utf-8") as f:
            cache = json.load(f)
        n_cache = len(cache) if isinstance(cache, dict) else 0
    except Exception:
        n_cache = 0

    print("=== Dry sector stats (kpmgfile.xlsx, pipeline normalize) ===")
    print(f"Company file: {company_file}")
    print(f"Total company rows (merged tabs): {len(companies)}")
    print(f"Sector non-blank: {int(sector_nonempty.sum())}")
    print(f"Sector blank (would use inference if vocab non-empty): {n_blank}")
    print(f"Closed vocabulary size (distinct non-blank Sector): {len(vocab)}")
    print(f"Sector inference cache entries ({SECTOR_INFERENCE_CACHE_FILE}): {n_cache}")

    optional = [
        "industry_field",
        "business_unit_field",
        "hq_field",
        "country_field",
        "website_field",
    ]
    populated = {
        oc: int(companies[oc].astype(str).map(safe_str).ne("").sum())
        for oc in optional
        if oc in companies.columns
    }
    print(f"Optional inference-context columns (non-empty rows): {populated}")

    if n_blank:
        print("\nFirst 15 blank-Sector company_name values:")
        mask = ~sector_nonempty
        for _, r in companies.loc[mask, ["company_name"]].head(15).iterrows():
            print(f"  - {safe_str(r['company_name'])[:120]}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN  (8-step pipeline)
# ─────────────────────────────────────────────────────────────────────────────
def main():
    load_project_dotenv()

    if len(sys.argv) > 1:
        cmd = sys.argv[1].lower()
        if cmd in ("--help", "-h"):
            _print_cli_help()
            return
        if cmd == "--resume-from-step8":
            resume_from_step8_only()
            return
        if cmd in ("--resume-export", "-r"):
            resume_export_only()
            return
        if cmd == "--dry-sector-stats":
            run_dry_sector_stats()
            return

    api_key = os.getenv("OPENAI_API_KEY", "")
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY is required.")
    client = OpenAI(api_key=api_key)

    company_file = find_first_existing(COMPANY_FILE_CANDIDATES)
    opp_file     = find_first_existing(OPPORTUNITY_FILE_CANDIDATES)
    print(f"Company file    : {company_file}")
    print(f"Opportunity file: {opp_file}")
    if "opportunities_structured" in os.path.basename(opp_file).lower():
        print("  Structured opportunity schema detected (multi-signal triage inputs).")

    if os.path.basename(company_file) != "kpmgfile.xlsx":
        raise ValueError(
            "Configured to use kpmgfile.xlsx as company source. Found "
            f"{company_file!r} instead. Add kpmgfile.xlsx to the project directory."
        )

    companies     = concat_matching_excel_sheets(company_file, normalize_company_cols, "Companies")
    opportunities = concat_matching_excel_sheets(
        opp_file, normalize_opportunities_for_pipeline, "Opportunities",
    )

    comp_id_col = find_id_column(companies, "company")
    opp_id_col  = find_id_column(opportunities, "opportunity")
    if comp_id_col:
        companies["_company_id"] = pd.to_numeric(companies[comp_id_col], errors="coerce").fillna(0).astype(int)
    else:
        companies["_company_id"] = (companies.index + 1).astype(int)
    if opp_id_col:
        opportunities["_opp_id"] = pd.to_numeric(opportunities[opp_id_col], errors="coerce").fillna(0).astype(int)
    else:
        opportunities["_opp_id"] = (opportunities.index + 1).astype(int)

    companies["_key"]     = companies.index.map(lambda i: f"C{i}")
    opportunities["_key"] = opportunities.index.map(lambda i: f"O{i}")

    for oc in (
        "industry_field",
        "business_unit_field",
        "hq_field",
        "country_field",
        "website_field",
    ):
        if oc not in companies.columns:
            companies[oc] = ""

    sector_vocab = closed_sector_vocabulary(companies)
    adj_sector_allowed = set(sector_vocab)
    enrich_companies_missing_sectors(
        client,
        companies,
        vocabulary=sector_vocab,
        cache_path=SECTOR_INFERENCE_CACHE_FILE,
        call_json_fn=call_json,
        model=ONTOLOGY_MODEL,
    )

    print(f"Loaded {len(companies)} companies × {len(opportunities)} opportunities = "
          f"{len(companies)*len(opportunities)} candidate pairs (full Cartesian)")
    name_filled = (
        companies["company_name"].fillna("").astype(str).str.strip().ne("")
    ).sum()
    print(f"  Company Name non-empty: {int(name_filled)} / {len(companies)} rows")

    # ── STEP 1 — Preprocessing ────────────────────────────────────────────────
    print("\n[Step 1/8] Preprocessing — normalizing text...")
    companies["_profile_text"] = companies["company_profile"].astype(str).map(preprocess)
    companies["_sector_norm"]  = companies["_effective_sector"].astype(str).map(canonicalize_sector)
    opportunities["_desc_text"]   = opportunities["What is the opportunity description?"].astype(str).map(preprocess)
    opportunities["_sector_norm"] = opportunities["Sector"].astype(str).map(canonicalize_sector)

    # ── STEP 2 — Sector Ontology Expansion ────────────────────────────────────
    print("[Step 2/8] Sector ontology expansion...")
    all_sectors = list(set(
        companies["_effective_sector"].astype(str).map(safe_str).tolist()
        + opportunities["Sector"].astype(str).map(safe_str).tolist()
    ))
    expanded = build_ontology(client, all_sectors, ONTOLOGY_CACHE_FILE)
    print(f"  Ontology built for {len(expanded)} unique sectors")

    # ── ENTITY EXTRACTION (feeds Step 5 product embedding) ────────────────────
    print("[Step 4a/8] Extracting product/capability entities (parallel + disk cache)...")
    opp_items = [
        {
            "id":          row["_key"],
            "name":        safe_str(row["What is the opportunity name?"]),
            "sector":      safe_str(row["Sector"]),
            "description": safe_str(row["What is the opportunity description?"]),
            "highlights":  safe_str(row.get("What are the investment highlights?", "")),
            "value_prop":  safe_str(row.get("What is the value proposition of this opportunity?", "")),
            "demand":      safe_str(row.get("What are the key demand drivers?", "")),
            "materials":   safe_str(row.get("What materials are involved or required in the project?", "")),
        }
        for _, row in opportunities.iterrows()
    ]

    comp_items = [
        {
            "id":               row["_key"],
            "company_name":     safe_str(row["company_name"]),
            "sector":           safe_str(row["_effective_sector"]),
            "profile":          safe_str(row["company_profile"]),
            "product_services": safe_str(row["product_services"]),
        }
        for _, row in companies.iterrows()
    ]

    ent_cache = load_entity_extraction_cache()
    for _ek in companies.loc[companies["sector_was_inferred"].astype(bool), "_key"]:
        ent_cache["company"].pop(_ek, None)
    populate_entity_texts_cached_parallel(
        client, opp_items, "opportunity", ent_cache["opportunity"], opp_entity_text,
    )
    populate_entity_texts_cached_parallel(
        client, comp_items, "company", ent_cache["company"], comp_entity_text,
    )
    save_entity_extraction_cache(ent_cache)
    opp_ent_raw = ent_cache["opportunity"]
    comp_ent_raw = ent_cache["company"]

    companies["_product_text"] = companies["_key"].map(
        lambda k: comp_entity_text(comp_ent_raw.get(k, {})),
    )
    opportunities["_product_text"] = opportunities["_key"].map(
        lambda k: opp_entity_text(opp_ent_raw.get(k, {})),
    )

    struct_texts = []
    for _, row in opportunities.iterrows():
        caps = parse_json_list(row.get("required_capabilities", "[]"))
        ins = parse_json_list(row.get("required_inputs", "[]"))
        blob = preprocess(" ".join(caps + ins)) if (caps or ins) else ""
        struct_texts.append(blob or preprocess(safe_str(row.get("What is the opportunity description?", ""))) or "n/a")
    opportunities["_struct_emb_text"] = struct_texts

    # ── STEP 4 — Semantic Embedding & Similarity ──────────────────────────────
    print(f"[Step 4/8] Building embeddings ({EMBEDDING_MODEL}, 1536-dim)...")
    use_openai = True
    try:
        companies["emb_profile"]      = embed_batch(client, companies["_profile_text"].tolist(), EMBEDDING_MODEL)
        companies["emb_product"]      = embed_batch(client, companies["_product_text"].tolist(), EMBEDDING_MODEL)
        opportunities["emb_desc"]     = embed_batch(client, opportunities["_desc_text"].tolist(), EMBEDDING_MODEL)
        opportunities["emb_product"]  = embed_batch(client, opportunities["_product_text"].tolist(), EMBEDDING_MODEL)
        opportunities["emb_struct"]   = embed_batch(
            client, opportunities["_struct_emb_text"].tolist(), EMBEDDING_MODEL,
        )
    except Exception as e:
        print(f"  OpenAI embedding failed ({e}). Using TF-IDF fallback.")
        use_openai = False
        all_texts = (
            companies["_profile_text"].tolist() + companies["_product_text"].tolist()
            + opportunities["_desc_text"].tolist() + opportunities["_product_text"].tolist()
            + opportunities["_struct_emb_text"].tolist()
        )
        tfidf = TfidfVectorizer(ngram_range=(1, 2), min_df=1)
        tfidf.fit(all_texts)
        companies["emb_profile"]      = list(tfidf.transform(companies["_profile_text"]).toarray())
        companies["emb_product"]      = list(tfidf.transform(companies["_product_text"]).toarray())
        opportunities["emb_desc"]     = list(tfidf.transform(opportunities["_desc_text"]).toarray())
        opportunities["emb_product"]  = list(tfidf.transform(opportunities["_product_text"]).toarray())
        opportunities["emb_struct"]   = list(tfidf.transform(opportunities["_struct_emb_text"]).toarray())

    # ── Multi-signal triage (replaces legacy sector-only gate + SOFT_MATCH_THRESHOLD routing)
    print("[Steps 3+5/8] Multi-signal triage routing…")
    rows: List[dict] = []
    bucket = Counter()
    n_to_gpt = 0
    n_filtered = 0

    opp_meta: List[dict] = []
    for _, row in opportunities.iterrows():
        raw_adj = parse_json_list(row.get("adjacent_value_chain_sectors", "[]"))
        adj_resolved = {r for x in raw_adj if (r := resolve_adjacent_label(x, adj_sector_allowed))}
        opp_meta.append({
            "adj": adj_resolved,
            "cap_kw": parse_json_list(row.get("capability_keywords", "[]")),
            "req_in": parse_json_list(row.get("required_inputs", "[]")),
        })

    comp_prof_m = np.asarray(companies["emb_profile"].tolist(), dtype=np.float64)
    comp_prod_m = np.asarray(companies["emb_product"].tolist(), dtype=np.float64)
    opp_desc_m = np.asarray(opportunities["emb_desc"].tolist(), dtype=np.float64)
    opp_prod_m = np.asarray(opportunities["emb_product"].tolist(), dtype=np.float64)
    opp_struct_m = np.asarray(opportunities["emb_struct"].tolist(), dtype=np.float64)
    sim_profile = cosine_similarity_matrix(comp_prof_m, opp_desc_m)
    sim_product = cosine_similarity_matrix(comp_prod_m, opp_prod_m)
    sim_struct_product = cosine_similarity_matrix(comp_prod_m, opp_struct_m)

    for oi, (_, opp) in enumerate(opportunities.iterrows()):
        opp_sector_raw = safe_str(opp["Sector"])
        meta = opp_meta[oi]
        for ci, (_, comp) in enumerate(companies.iterrows()):
            comp_sector_effective = safe_str(comp["_effective_sector"])

            profile_sim = round(float(sim_profile[ci, oi]), 3)
            product_sim = round(float(sim_product[ci, oi]), 3)
            combined = max(profile_sim, product_sim)

            sector_overlap_bool = sectors_overlap(comp_sector_effective, opp_sector_raw, expanded)
            sector_ov_int = 1 if sector_overlap_bool else 0
            val_chain_adj = comp_sector_effective in meta["adj"]
            corp_for_kw = safe_str(comp["product_services"]) + " " + safe_str(comp["company_profile"])
            cap_kw_hits = keyword_match_count(meta["cap_kw"], corp_for_kw)
            input_hits = keyword_match_count(meta["req_in"], safe_str(comp["product_services"]))
            sig_prof_routing = profile_sim
            sig_prod_routing = round(float(sim_struct_product[ci, oi]), 3)

            t_sec = sector_ov_int == 1
            t_vc = val_chain_adj
            t_kw = cap_kw_hits >= TRIAGE_MIN_CAPABILITY_KEYWORD_HITS
            t_in = input_hits >= TRIAGE_MIN_INPUT_OVERLAP_HITS
            t_ps = sig_prod_routing >= TRIAGE_MIN_SIGNAL_PRODUCT_SIM
            t_pf = sig_prof_routing >= TRIAGE_MIN_SIGNAL_PROFILE_SIM
            send_to_gpt = t_sec or t_vc or t_kw or t_in or t_ps or t_pf

            pass_labels: List[str] = []
            if t_sec:
                pass_labels.append("sector_overlap")
            if t_vc:
                pass_labels.append("value_chain_adjacent")
            if t_kw:
                pass_labels.append(f"capability_keywords:{cap_kw_hits}")
            if t_in:
                pass_labels.append(f"input_overlap:{input_hits}")
            if t_ps:
                pass_labels.append(f"signal_product_sim:{sig_prod_routing:.3f}")
            if t_pf:
                pass_labels.append(f"signal_profile_sim:{sig_prof_routing:.3f}")

            if send_to_gpt:
                triage_reason = format_pass_reason(pass_labels)
                if t_sec:
                    pass_reason = "sector_overlap"
                elif t_vc:
                    pass_reason = "value_chain_adjacent"
                elif t_kw:
                    pass_reason = "capability_keywords"
                elif t_in:
                    pass_reason = "input_overlap"
                elif t_ps:
                    pass_reason = "signal_product_similarity"
                elif t_pf:
                    pass_reason = "signal_profile_similarity"
                else:
                    pass_reason = "triage_pass"
                n_to_gpt += 1
                if t_sec:
                    bucket["sector_overlap"] += 1
                if t_vc:
                    bucket["value_chain_adjacent"] += 1
                if t_kw:
                    bucket["capability_keywords"] += 1
                if t_in:
                    bucket["input_overlap"] += 1
                if t_ps:
                    bucket["signal_product_similarity"] += 1
                if t_pf:
                    bucket["signal_profile_similarity"] += 1
            else:
                triage_reason = format_fail_reason(
                    overlap=sector_ov_int,
                    adj=t_vc,
                    ck=cap_kw_hits,
                    inp=input_hits,
                    sig_p=float(sig_prod_routing),
                    sig_pf=float(sig_prof_routing),
                )
                pass_reason = "filtered"
                n_filtered += 1

            rows.append({
                "_pair_key":  f"{opp['_key']}_{comp['_key']}",
                "_opp_key":   opp["_key"],
                "_comp_key":  comp["_key"],
                "_pass_reason": pass_reason,
                "_send_to_gpt": send_to_gpt,
                "_combined":   float(combined),
                "companyId":     int(comp["_company_id"]),
                "opportunityId": int(opp["_opp_id"]),
                "company_name":             safe_str(comp["company_name"]),
                "opportunity_name":         safe_str(opp[OPP_NAME_COL]),
                "company_profile":          safe_str(comp["company_profile"]),
                "opportunity_description":  safe_str(opp[OPP_DESC_COL]),
                "company_sector":     safe_str(comp["Sector"]) or comp_sector_effective,
                "opportunity_sector": opp_sector_raw,
                "sector_similarity":        float(combined),
                "_sector_ontology_overlap": sector_ov_int,
                "profile_similarity":       float(profile_sim),
                "product_similarity":       float(product_sim),
                "signal_sector_overlap":    int(sector_ov_int),
                "signal_value_chain_adjacent": bool(t_vc),
                "signal_capability_keyword_count": int(cap_kw_hits),
                "signal_input_overlap_count": int(input_hits),
                "signal_product_similarity": float(sig_prod_routing),
                "signal_profile_similarity": float(sig_prof_routing),
                "triage_pass": bool(send_to_gpt),
                "triage_reason": triage_reason,
                "ai_score":       0,
                "ai_decision":    "",
                "ai_explanation": "",
                "ai_insight":     "",
                "suggested_plan": [],
                "match_reason":   [],
                "sector_was_inferred":      bool(comp.get("sector_was_inferred", False)),
                "sector_inference_source": safe_str(comp.get("sector_inference_source", "none")),
                "sector_inference_evidence": safe_str(comp.get("sector_inference_evidence", "")),
            })

    df = pd.DataFrame(rows)
    total_pairs = len(companies) * len(opportunities)
    print(f"  Total pairs: {total_pairs}; sent to GPT: {n_to_gpt}; filtered by triage: {n_filtered}")
    print(f"  Triage signal hit counts (non-exclusive; pair can count in several): {dict(bucket)}")

    # ── STEP 6 — GPT validation on eligible pairs ─────────────────────────────
    print("[Step 6/8] GPT validation (analyst-grade Yes/No + narrative)...")
    eligible = df[df["_send_to_gpt"]].copy()

    def collect_gpt_updates_for_opportunity(
        opp_key: str,
        idxs: List[int],
    ) -> List[Tuple[int, Dict[str, object]]]:
        """Read-only on ``df``; returns row updates applied on the main thread."""
        opp = opportunities[opportunities["_key"] == opp_key].iloc[0]
        oent = opp_ent_raw.get(opp_key, {})
        opp_context = {
            "opportunity_name": safe_str(opp["What is the opportunity name?"]),
            "opportunity_sector": safe_str(opp["Sector"]),
            "opportunity_description": safe_str(opp["What is the opportunity description?"]),
            "investment_highlights": safe_str(opp.get("What are the investment highlights?", "")),
            "value_proposition": safe_str(opp.get("What is the value proposition of this opportunity?", "")),
            "demand_drivers": safe_str(opp.get("What are the key demand drivers?", "")),
            "materials": safe_str(opp.get("What materials are involved or required in the project?", "")),
            "required_products":  to_list(oent.get("required_products")),
            "required_services":  to_list(oent.get("required_services")),
            "required_materials": to_list(oent.get("required_materials")),
            "required_capabilities": to_list(oent.get("required_capabilities")),
        }

        patch: List[Tuple[int, Dict[str, object]]] = []

        for i in range(0, len(idxs), DECISION_BATCH):
            bidx = idxs[i : i + DECISION_BATCH]
            candidates = []
            for idx in bidx:
                comp_key = df.at[idx, "_comp_key"]
                comp = companies[companies["_key"] == comp_key].iloc[0]
                cent = comp_ent_raw.get(comp_key, {})
                candidates.append({
                    "candidate_id":    df.at[idx, "_pair_key"],
                    "company_name":    safe_str(comp["company_name"]),
                    "company_sector":  safe_str(comp["_effective_sector"]),
                    "company_profile": safe_str(comp["company_profile"])[:1800],
                    "product_services": safe_str(comp["product_services"])[:1800],
                    "actual_products":  to_list(cent.get("actual_products")),
                    "actual_services":  to_list(cent.get("actual_services")),
                    "capabilities":     to_list(cent.get("capabilities")),
                    "match_path": safe_str(df.at[idx, "triage_reason"])[:520],
                    "ranking_signals": {
                        "ontology_sector_overlap": int(df.at[idx, "_sector_ontology_overlap"]),
                        "profile_similarity": float(df.at[idx, "profile_similarity"]),
                        "product_similarity": float(df.at[idx, "product_similarity"]),
                        "combined_max": float(df.at[idx, "_combined"]),
                    },
                })

            try:
                results_map = gpt_decide_batch(client, opp_context, candidates)
            except Exception as e:
                print(f"    GPT decision failed [{opp_key}]: {e}")
                results_map = {}

            for idx in bidx:
                cid = df.at[idx, "_pair_key"]
                raw = results_map.get(cid, {"candidate_id": cid, "ai_decision": "No"})
                r = safe_decision(raw, cid)

                patch.append((idx, {
                    "ai_decision": r["ai_decision"],
                    "ai_score": 1 if r["ai_decision"] == "Yes" else 0,
                    "ai_explanation": r["ai_explanation"],
                    "ai_insight": r["ai_insight"],
                    "suggested_plan": r["suggested_plan"],
                    "match_reason": r["match_reason"],
                }))

        return patch

    def apply_row_updates(updates: List[Tuple[int, Dict[str, object]]]) -> None:
        for idx, vals in updates:
            for col, val in vals.items():
                df.at[idx, col] = val

    opp_groups = list(eligible.groupby("_opp_key", sort=False))
    n_grp = len(opp_groups)

    if PARALLEL_OPP_GPT < 2 or n_grp == 1:
        for opp_key, group in tqdm(opp_groups, total=n_grp, desc="  GPT by opportunity"):
            apply_row_updates(
                collect_gpt_updates_for_opportunity(opp_key, group.index.tolist()),
            )
    else:
        workers = min(PARALLEL_OPP_GPT, n_grp)
        with ThreadPoolExecutor(max_workers=workers) as pool:
            futs = {
                pool.submit(
                    collect_gpt_updates_for_opportunity,
                    opp_key,
                    group.index.tolist(),
                ): opp_key
                for opp_key, group in opp_groups
            }
            for fut in tqdm(as_completed(futs), total=len(futs), desc="  GPT by opportunity"):
                ok = futs[fut]
                try:
                    apply_row_updates(fut.result())
                except Exception as e:
                    print(f"  GPT opportunity worker failed [{ok}]: {e}")

    stamp_sector_gate_filtered(df)

    os.makedirs(os.path.dirname(CHECKPOINT_POST_STEP6) or ".", exist_ok=True)
    df.to_pickle(CHECKPOINT_POST_STEP6)
    print(f"  Post–Step 6 checkpoint: {CHECKPOINT_POST_STEP6} (resume: --resume-from-step8)")
    finalize_ranking_and_export(df, companies_summary=companies)


if __name__ == "__main__":
    main()
